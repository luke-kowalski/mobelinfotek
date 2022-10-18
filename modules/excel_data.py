from openpyxl import load_workbook
from modules import ms_sql_database, settings, files_handling
import os


def check_excel_data(file_name):
    try:
        wb = load_workbook(filename=os.path.join(settings.EXCEL_SOURCE_PATH, file_name))
        sheet_names = wb.sheetnames
        if validate_sheet_name(sheet_names) is True:
            sheet = wb[settings.EXCEL_SHEET_NAME]
            if sheet.max_column >= 5:
                print(f"OK: Working with data from the Excel file: {file_name}")
                delete_existed_entries_in_ms_sql(sheet)
                insert_excel_data_to_ms_sql(sheet)
                files_handling.move_excel_file_to_archive_dir(file_name)
            else:
                print(f"ERROR: Column number is less then 5! File name: {file_name}")
        else:
            print(f"ERROR: There is no sheet name {settings.EXCEL_SHEET_NAME} in the Excel file: {file_name} !")
    except PermissionError as e:
        print(f"ERROR: PLEASE CLOSE EXCEL FILE! : {e}")


def validate_sheet_name(sheet_names):
    for sheet in sheet_names:
        if sheet == settings.EXCEL_SHEET_NAME:
            return True


def delete_existed_entries_in_ms_sql(sheet):
    data = sheet
    orders_name_list = set()
    for row in data.iter_rows(min_row=int(settings.EXCEL_START_ROW_NUMBER), values_only=True):
        imos_order_name = row[3]
        orders_name_list.add(imos_order_name)

    for order_name in orders_name_list:
        ms_sql_database.ms_sql_delete_orders(order_name)


def insert_excel_data_to_ms_sql(sheet):
    data = sheet
    inserted_row = 0
    for row in data.iter_rows(min_row=int(settings.EXCEL_START_ROW_NUMBER), values_only=True):
        building = row[0]
        building_floor = row[1]
        building_unit = row[2]
        imos_order_name = row[3]
        furniture_type = row[4]
        ms_sql_database.ms_sql_insert_data_to_database(building, building_floor, building_unit, imos_order_name, furniture_type)
        inserted_row += 1
    print(f"OK: Inserted rows: {inserted_row}")
